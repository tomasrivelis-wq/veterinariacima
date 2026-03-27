import streamlit as st
import pandas as pd
import numpy as np
from statsmodels.tsa.holtwinters import ExponentialSmoothing
from io import BytesIO
import warnings
warnings.filterwarnings('ignore')

# ─────────────────────────────────────────────
#  CONFIGURACIÓN DE PÁGINA
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="VetForecast 2026",
    page_icon="🐾",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────
#  ESTILOS
# ─────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=Space+Mono:wght@400;700&display=swap');

html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
}

/* Sidebar */
section[data-testid="stSidebar"] {
    background: linear-gradient(160deg, #0f1b2d 0%, #1a2e4a 100%);
    border-right: 1px solid #2a4a6b;
}
section[data-testid="stSidebar"] * { color: #c9dff2 !important; }
section[data-testid="stSidebar"] h1, 
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 {
    color: #7ec8e3 !important;
    font-family: 'Space Mono', monospace !important;
    font-size: 0.85rem !important;
    letter-spacing: 0.1em;
    text-transform: uppercase;
}

/* Main background */
.main { background-color: #f5f7fa; }

/* Header */
.hero-header {
    background: linear-gradient(135deg, #0f1b2d 0%, #1e3a5f 50%, #0f1b2d 100%);
    border-radius: 16px;
    padding: 2.5rem 2rem;
    margin-bottom: 2rem;
    border: 1px solid #2a4a6b;
    position: relative;
    overflow: hidden;
}
.hero-header::before {
    content: '';
    position: absolute;
    top: -50%;
    right: -10%;
    width: 300px;
    height: 300px;
    border-radius: 50%;
    background: radial-gradient(circle, rgba(126,200,227,0.08) 0%, transparent 70%);
}
.hero-title {
    font-family: 'Space Mono', monospace;
    font-size: 2.2rem;
    font-weight: 700;
    color: #ffffff;
    margin: 0;
    line-height: 1.2;
}
.hero-subtitle {
    font-size: 1rem;
    color: #7ec8e3;
    margin-top: 0.5rem;
    font-weight: 400;
}
.hero-badge {
    display: inline-block;
    background: rgba(126,200,227,0.15);
    border: 1px solid rgba(126,200,227,0.3);
    color: #7ec8e3;
    font-family: 'Space Mono', monospace;
    font-size: 0.7rem;
    padding: 0.25rem 0.75rem;
    border-radius: 20px;
    margin-bottom: 1rem;
    letter-spacing: 0.1em;
    text-transform: uppercase;
}

/* Cards */
.metric-card {
    background: white;
    border-radius: 12px;
    padding: 1.25rem 1.5rem;
    border: 1px solid #e2e8f0;
    box-shadow: 0 1px 3px rgba(0,0,0,0.05);
}
.metric-value {
    font-family: 'Space Mono', monospace;
    font-size: 2rem;
    font-weight: 700;
    color: #0f1b2d;
    line-height: 1;
}
.metric-label {
    font-size: 0.8rem;
    color: #6b7280;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    margin-top: 0.4rem;
    font-weight: 500;
}
.metric-sub {
    font-size: 0.75rem;
    color: #9ca3af;
    margin-top: 0.2rem;
}

/* Status pill */
.pill-activo {
    background: #d1fae5; color: #065f46;
    border-radius: 20px; padding: 0.15rem 0.6rem;
    font-size: 0.7rem; font-weight: 600;
    letter-spacing: 0.05em; text-transform: uppercase;
}
.pill-inactivo {
    background: #fee2e2; color: #991b1b;
    border-radius: 20px; padding: 0.15rem 0.6rem;
    font-size: 0.7rem; font-weight: 600;
    letter-spacing: 0.05em; text-transform: uppercase;
}

/* Upload zone */
.upload-label {
    font-size: 0.75rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    color: #374151;
    margin-bottom: 0.3rem;
}

/* Info box */
.info-box {
    background: #eff6ff;
    border-left: 3px solid #3b82f6;
    border-radius: 0 8px 8px 0;
    padding: 0.75rem 1rem;
    font-size: 0.85rem;
    color: #1e40af;
    margin: 0.5rem 0;
}
.warn-box {
    background: #fffbeb;
    border-left: 3px solid #f59e0b;
    border-radius: 0 8px 8px 0;
    padding: 0.75rem 1rem;
    font-size: 0.85rem;
    color: #92400e;
    margin: 0.5rem 0;
}

/* Progress */
.step-label {
    font-family: 'Space Mono', monospace;
    font-size: 0.7rem;
    color: #6b7280;
    text-transform: uppercase;
    letter-spacing: 0.1em;
}

/* Download btn override */
div[data-testid="stDownloadButton"] > button {
    background: linear-gradient(135deg, #0f1b2d, #1e3a5f) !important;
    color: white !important;
    border: none !important;
    padding: 0.75rem 2rem !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    font-family: 'DM Sans', sans-serif !important;
    width: 100% !important;
    font-size: 1rem !important;
}

hr { border-color: #e5e7eb; margin: 1.5rem 0; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
#  FUNCIONES DE LÓGICA
# ─────────────────────────────────────────────

FAMILIAS = ['MEDICAMENTOS FARMACIA', 'ACCESORIOS', 'BALANCEADOS']
FECHA_CORTE_RECENCIA = pd.Timestamp('2024-07-01')
MIN_VENTAS = 5
MESES_2026 = [
    'Ene-2026','Feb-2026','Mar-2026','Abr-2026','May-2026','Jun-2026',
    'Jul-2026','Ago-2026','Sep-2026','Oct-2026','Nov-2026','Dic-2026'
]


def leer_excel(uploaded_file):
    """Lee un archivo Excel subido, retorna DataFrame o None."""
    try:
        df = pd.read_excel(uploaded_file, engine='openpyxl')
        return df
    except Exception as e:
        st.error(f"❌ Error al leer el archivo: {e}")
        return None


def validar_columnas(df, columnas_requeridas, nombre_archivo):
    """Verifica que el DataFrame tenga las columnas necesarias."""
    faltantes = [c for c in columnas_requeridas if c not in df.columns]
    if faltantes:
        st.error(f"❌ **{nombre_archivo}** no tiene las columnas: `{', '.join(faltantes)}`")
        st.info(f"Columnas encontradas: `{', '.join(df.columns.tolist())}`")
        return False
    return True


def preprocesar_ventas(df_ventas, df_maestro):
    """Sanea y filtra el DataFrame de ventas."""
    df_ventas = df_ventas.copy()
    df_maestro = df_maestro.copy()

    # Normalizar claves
    df_ventas['Artículo'] = df_ventas['Artículo'].astype(str).str.strip()
    df_maestro['Artículo'] = df_maestro['Artículo'].astype(str).str.strip()

    # Parsear fecha
    df_ventas['Fecha'] = pd.to_datetime(df_ventas['Fecha'], dayfirst=True, errors='coerce')
    df_ventas = df_ventas.dropna(subset=['Fecha'])

    # Filtro de familias
    df_ventas['Descrip. familia'] = df_ventas['Descrip. familia'].astype(str).str.strip().str.upper()
    df_f = df_ventas[df_ventas['Descrip. familia'].isin([f.upper() for f in FAMILIAS])].copy()

    if df_f.empty:
        return df_f, pd.Index([]), df_maestro

    # Cantidad positiva
    df_f['Cantidad'] = pd.to_numeric(df_f['Cantidad'], errors='coerce').fillna(0)
    df_f = df_f[df_f['Cantidad'] > 0]

    # Clasificación ACTIVO / INACTIVO
    resumen = df_f.groupby('Artículo').agg(
        total=('Cantidad', 'sum'),
        ultima=('Fecha', 'max')
    )
    activos = resumen[
        (resumen['total'] >= MIN_VENTAS) & (resumen['ultima'] >= FECHA_CORTE_RECENCIA)
    ].index

    return df_f, activos, df_maestro


def calcular_forecast_estacional(serie: pd.Series) -> list:
    """Holt-Winters con corrección anti-meseta."""
    try:
        if len(serie) < 24:
            # Rellenar con ceros si la serie es corta para cumplir seasonal_periods=12
            full_index = pd.date_range(
                end=serie.index[-1], periods=24, freq='MS'
            )
            serie = serie.reindex(full_index, fill_value=0)

        modelo = ExponentialSmoothing(
            serie,
            trend='add',
            seasonal='add',
            seasonal_periods=12
        ).fit(
            smoothing_level=0.75,
            optimized=False
        )
        pred = modelo.forecast(12).values

        # Clip negativos
        pred = np.clip(pred, 0, None)

        # Anti-meseta: si todos los valores son iguales
        if len(set(np.round(pred, 2))) == 1:
            base = serie.tail(3).mean() if serie.tail(3).mean() > 0 else max(serie.mean(), 0.1)
            pred = np.array([base * (1 + i * 0.02) for i in range(1, 13)])

        return pred.tolist()

    except Exception:
        base = max(serie.tail(6).mean(), 0)
        return [base] * 12


def construir_serie_mensual(df_articulo: pd.DataFrame) -> pd.Series:
    """Agrega ventas por mes y crea serie temporal completa."""
    df_articulo = df_articulo.copy()
    df_articulo['Mes'] = df_articulo['Fecha'].dt.to_period('M').dt.to_timestamp()
    mensual = df_articulo.groupby('Mes')['Cantidad'].sum()

    if mensual.empty:
        return pd.Series([], dtype=float)

    full_range = pd.date_range(
        start=mensual.index.min(),
        end=mensual.index.max(),
        freq='MS'
    )
    mensual = mensual.reindex(full_range, fill_value=0)
    return mensual


def procesar_forecasts(df_ventas_filtrado, df_maestro, activos, progress_bar, status_text):
    """Itera sobre todos los artículos del maestro y genera pronósticos."""
    todos_codigos = df_maestro['Artículo'].unique()
    total = len(todos_codigos)
    resultados = []

    for i, codigo in enumerate(todos_codigos):
        # Actualizar progreso
        progress_bar.progress((i + 1) / total)
        status_text.markdown(
            f'<span class="step-label">Procesando artículo {i+1} / {total} → `{codigo}`</span>',
            unsafe_allow_html=True
        )

        es_activo = codigo in activos
        estado = "ACTIVO" if es_activo else "INACTIVO"

        # Descripción del producto (tomar la más reciente)
        descrip = ""
        familia = ""
        df_art = df_ventas_filtrado[df_ventas_filtrado['Artículo'] == codigo]

        if not df_art.empty:
            descrip_col = next((c for c in df_art.columns if 'descrip' in c.lower() and 'famil' not in c.lower()), None)
            if descrip_col:
                descrip = df_art[descrip_col].iloc[-1]
            fam_col = next((c for c in df_art.columns if 'famil' in c.lower()), None)
            if fam_col:
                familia = df_art[fam_col].iloc[-1]

        if not es_activo:
            fila = {
                'Código': codigo,
                'Descripción': descrip,
                'Familia': familia,
                'Estado': estado,
            }
            fila.update({m: 0 for m in MESES_2026})
            resultados.append(fila)
            continue

        # Calcular forecast
        serie = construir_serie_mensual(df_art)

        if serie.empty or len(serie) < 2:
            forecast = [0] * 12
        else:
            forecast = calcular_forecast_estacional(serie)

        # Redondear a enteros (unidades)
        forecast_int = [max(0, round(v)) for v in forecast]

        fila = {
            'Código': codigo,
            'Descripción': descrip,
            'Familia': familia,
            'Estado': estado,
        }
        fila.update({m: v for m, v in zip(MESES_2026, forecast_int)})
        resultados.append(fila)

    return pd.DataFrame(resultados)


def generar_excel(df_resultado: pd.DataFrame) -> bytes:
    """Convierte el DataFrame de resultados a bytes de Excel."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_resultado.to_excel(writer, index=False, sheet_name='Forecast 2026')

        # Formateo básico
        ws = writer.sheets['Forecast 2026']
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Header style
        header_fill = PatternFill(start_color="0F1B2D", end_color="0F1B2D", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        thin = Side(style='thin', color='E2E8F0')
        border = Border(bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Colores por estado
        green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        red_fill   = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

        estado_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'Estado':
                estado_col = idx
                break

        if estado_col:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                val = row[estado_col - 1].value
                fill = green_fill if val == 'ACTIVO' else red_fill
                row[estado_col - 1].fill = fill

        # Ajustar anchos
        col_widths = {'Código': 14, 'Descripción': 35, 'Familia': 22, 'Estado': 12}
        for i, cell in enumerate(ws[1], 1):
            col_letter = get_column_letter(i)
            w = col_widths.get(cell.value, 12)
            ws.column_dimensions[col_letter].width = w

        ws.freeze_panes = 'E2'

    return output.getvalue()


# ─────────────────────────────────────────────
#  SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 🐾 VetForecast")
    st.markdown("---")
    st.markdown("#### Parámetros del modelo")
    st.markdown(f"""
    - **Modelo:** Holt-Winters (Triple ES)  
    - **Tendencia:** Aditiva  
    - **Estacionalidad:** Aditiva  
    - **Períodos estacionales:** 12  
    - **Alpha (smoothing):** 0.75  
    """)
    st.markdown("---")
    st.markdown("#### Criterios de actividad")
    st.markdown(f"""
    - **Mín. ventas históricas:** {MIN_VENTAS}  
    - **Recencia mínima:** Julio 2024  
    - **Familias:** {', '.join(FAMILIAS)}  
    """)
    st.markdown("---")
    st.markdown("#### Sobre el pronóstico")
    st.markdown("""
    <div class="info-box">
    Los productos INACTIVOS siempre aparecen en el output con pronóstico = 0, manteniendo la visibilidad total del catálogo.
    </div>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────
#  HEADER
# ─────────────────────────────────────────────
st.markdown("""
<div class="hero-header">
    <div class="hero-badge">🔬 Forecasting Engine v1.0</div>
    <div class="hero-title">🐾 VetForecast 2026</div>
    <div class="hero-subtitle">Pronóstico de demanda mensual para Pet Shop & Veterinaria · Holt-Winters Triple Exponential Smoothing</div>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
#  CARGA DE ARCHIVOS
# ─────────────────────────────────────────────
col1, col2 = st.columns(2)

with col1:
    st.markdown('<div class="upload-label">📦 1. Historial de Ventas</div>', unsafe_allow_html=True)
    ventas_file = st.file_uploader(
        "Ventas.xlsx",
        type=['xlsx'],
        key='ventas',
        label_visibility='collapsed'
    )
    if ventas_file:
        st.success(f"✅ `{ventas_file.name}` cargado ({ventas_file.size // 1024} KB)")

with col2:
    st.markdown('<div class="upload-label">📋 2. Maestro de Productos</div>', unsafe_allow_html=True)
    maestro_file = st.file_uploader(
        "Maestro.xlsx",
        type=['xlsx'],
        key='maestro',
        label_visibility='collapsed'
    )
    if maestro_file:
        st.success(f"✅ `{maestro_file.name}` cargado ({maestro_file.size // 1024} KB)")

st.markdown("---")

# ─────────────────────────────────────────────
#  PROCESAMIENTO
# ─────────────────────────────────────────────
if ventas_file and maestro_file:
    if st.button("🚀 Generar Forecast 2026", use_container_width=True, type="primary"):
        with st.spinner("Iniciando procesamiento..."):

            # 1. Leer archivos
            df_ventas  = leer_excel(ventas_file)
            df_maestro = leer_excel(maestro_file)

            if df_ventas is None or df_maestro is None:
                st.stop()

            # 2. Validar columnas
            cols_ventas  = ['Artículo', 'Fecha', 'Cantidad', 'Descrip. familia']
            cols_maestro = ['Artículo']

            if not validar_columnas(df_ventas, cols_ventas, "Ventas.xlsx"):
                st.stop()
            if not validar_columnas(df_maestro, cols_maestro, "Maestro.xlsx"):
                st.stop()

            # 3. Preprocesar
            df_filtrado, activos, df_maestro = preprocesar_ventas(df_ventas, df_maestro)

            if df_filtrado.empty:
                st.warning("⚠️ No se encontraron registros con las familias requeridas.")
                st.stop()

            # 4. Métricas rápidas
            n_maestro  = df_maestro['Artículo'].nunique()
            n_activos  = len(activos)
            n_inactivos = n_maestro - n_activos
            n_registros = len(df_filtrado)

            m1, m2, m3, m4 = st.columns(4)
            with m1:
                st.markdown(f"""
                <div class="metric-card">
                    <div class="metric-value">{n_maestro:,}</div>
                    <div class="metric-label">Total productos</div>
                    <div class="metric-sub">En el maestro</div>
                </div>""", unsafe_allow_html=True)
            with m2:
                st.markdown(f"""
                <div class="metric-card">
                    <div class="metric-value" style="color:#065f46">{n_activos:,}</div>
                    <div class="metric-label">Productos activos</div>
                    <div class="metric-sub">Con forecast estacional</div>
                </div>""", unsafe_allow_html=True)
            with m3:
                st.markdown(f"""
                <div class="metric-card">
                    <div class="metric-value" style="color:#991b1b">{n_inactivos:,}</div>
                    <div class="metric-label">Productos inactivos</div>
                    <div class="metric-sub">Forecast = 0</div>
                </div>""", unsafe_allow_html=True)
            with m4:
                st.markdown(f"""
                <div class="metric-card">
                    <div class="metric-value">{n_registros:,}</div>
                    <div class="metric-label">Registros procesados</div>
                    <div class="metric-sub">Transacciones válidas</div>
                </div>""", unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            # 5. Forecast
            st.markdown("**⚙️ Calculando pronósticos...**")
            progress_bar = st.progress(0)
            status_text  = st.empty()

            df_resultado = procesar_forecasts(
                df_filtrado, df_maestro, activos, progress_bar, status_text
            )

            progress_bar.progress(1.0)
            status_text.markdown(
                '<span class="step-label">✅ Procesamiento completado</span>',
                unsafe_allow_html=True
            )

            # 6. Preview
            st.markdown("---")
            st.markdown("#### 🔍 Vista previa del resultado")

            col_preview, col_stats = st.columns([3, 1])
            with col_preview:
                st.dataframe(
                    df_resultado.head(20),
                    use_container_width=True,
                    height=350
                )
            with col_stats:
                total_forecast = df_resultado[MESES_2026].sum().sum()
                top_familia = (
                    df_resultado[df_resultado['Estado'] == 'ACTIVO']
                    .groupby('Familia')[MESES_2026].sum().sum(axis=1)
                    .idxmax() if n_activos > 0 else "—"
                )
                st.markdown(f"""
                <div class="metric-card" style="margin-bottom:1rem">
                    <div class="metric-value" style="font-size:1.5rem">{int(total_forecast):,}</div>
                    <div class="metric-label">Unidades forecast total 2026</div>
                </div>
                <div class="metric-card">
                    <div class="metric-value" style="font-size:1rem; color:#1e3a5f">{top_familia}</div>
                    <div class="metric-label">Familia con mayor volumen</div>
                </div>
                """, unsafe_allow_html=True)

            # 7. Descarga
            st.markdown("---")
            excel_bytes = generar_excel(df_resultado)

            dl_col, _ = st.columns([1, 2])
            with dl_col:
                st.download_button(
                    label="⬇️  Descargar Forecast_2026.xlsx",
                    data=excel_bytes,
                    file_name="Forecast_2026.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            st.markdown("""
            <div class="info-box" style="margin-top:1rem">
            ✅ El archivo incluye todos los productos del maestro. Los activos con pronóstico Holt-Winters y los inactivos con 0.
            </div>
            """, unsafe_allow_html=True)

elif ventas_file or maestro_file:
    st.markdown("""
    <div class="warn-box">
    ⚠️ Cargá <strong>ambos archivos</strong> para habilitar el procesamiento.
    </div>
    """, unsafe_allow_html=True)
else:
    st.markdown("""
    <div class="info-box">
    👆 Cargá los dos archivos (.xlsx) para comenzar. Podés usar los cargadores de arriba.
    </div>
    """, unsafe_allow_html=True)
