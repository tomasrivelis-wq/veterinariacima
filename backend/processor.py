"""
Roadmap Analytics — processor.py v4.0
======================================
v4.0:
  - get_forecast_filtros: cada artículo en la lista `articulos` incluye ahora
    los campos `familia` y `categoria`, necesarios para el auto-relleno
    en cascada del frontend (MOD 3: seleccionar artículo inyecta familia/categoría).
  - Sin cambios en la matemática de S_MIN, PEDIR ni en el motor Holt-Winters.
"""

from __future__ import annotations

import io
import os
import unicodedata
import warnings
from datetime import date
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill

try:
    from statsmodels.tsa.holtwinters import ExponentialSmoothing as _HW
    _HAS_STATSMODELS = True
except ImportError:
    _HAS_STATSMODELS = False

warnings.filterwarnings("ignore")

# ═════════════════════════════════════════════════════════════════════════════
# CONSTANTES — StockProcessor
# ═════════════════════════════════════════════════════════════════════════════
XLSM_SHEET      = "Planificación"
XLSM_COLS_IDX   = [0, 1, 2, 5, 6, 7, 53, 54]
XLSM_COLS_NAMES = [
    "CODIGO", "DESCRIPCION", "FAMILIA", "PROVEEDOR",
    "COMPRA_MIN", "LOTE", "S_MIN_REF", "PEDIR_REF",
]
PROV_EXCLUDE = {"", "0", "FALSO", "FALSE", "SIN COMPRAS", "#N/A", "NONE"}
SAFETY_MULTIPLIERS = {
    "2 semanas": 0.5, "3 semanas": 0.75, "1 mes": 1.0, "2 meses": 2.0,
}
DEFAULT_PARAMS = {
    "ACCESORIOS":             {"lead_time": 2, "stock_seguridad": "2 semanas"},
    "BALANCEADOS":            {"lead_time": 2, "stock_seguridad": "2 semanas"},
    "MEDICAMENTOS FARMACIA":  {"lead_time": 2, "stock_seguridad": "2 semanas"},
}
STATUS_SIN_STOCK  = "Sin Stock"
STATUS_FALTANTE   = "Faltante"
STATUS_SOBRESTOCK = "Sobrestock"
STATUS_NORMAL     = "Normal"
STATUS_COLORS = {
    STATUS_SIN_STOCK:  "FFC7CE",
    STATUS_FALTANTE:   "FFEB9C",
    STATUS_SOBRESTOCK: "BDD7EE",
    STATUS_NORMAL:     "C6EFCE",
}

# ═════════════════════════════════════════════════════════════════════════════
# CONSTANTES — ForecastEngine
# ═════════════════════════════════════════════════════════════════════════════
FC_FAMILIAS         = ["MEDICAMENTOS FARMACIA", "ACCESORIOS", "BALANCEADOS"]
FC_MIN_VENTAS       = 5
FC_MESES_RECENCIA   = 18
FC_SEASONAL_PERIODS = 12

# v3.5 — Parámetros Holt-Winters Amortiguado (Damped Trend)
FC_SMOOTHING_LEVEL  = 0.4
FC_SMOOTHING_TREND  = 0.2
FC_DAMPING_TREND    = 0.85


# ═════════════════════════════════════════════════════════════════════════════
# StockProcessor
# ═════════════════════════════════════════════════════════════════════════════
class StockProcessor:

    def __init__(self, xlsm_path: str, forecast_path: str):
        self.xlsm_path     = Path(xlsm_path)
        self.forecast_path = Path(forecast_path)
        self._master_cache: Optional[pd.DataFrame] = None
        self._active_codes_cache: Optional[set]    = None

    # ══════════════════════════════════════════════════════════════════════════
    # API PÚBLICA
    # ══════════════════════════════════════════════════════════════════════════

    def process(self, path_med, path_acc, path_bal,
                family_filter="Todas", params=None, debug_dir=None) -> pd.DataFrame:
        params = params or DEFAULT_PARAMS
        df     = self._full_calc(path_med, path_acc, path_bal, params, debug_dir)
        df     = self._apply_active_filter(df)
        df     = self._apply_filter(df, family_filter)
        result = (df[df["PEDIR"] > 0]
                  .sort_values(["PROVEEDOR", "FAMILIA", "CODIGO"])
                  .reset_index(drop=True))
        print(f"[PROC] PEDIR>0: {len(result)}  Total uds: {result['PEDIR'].sum():,}", flush=True)
        return result[["PROVEEDOR", "FAMILIA", "CODIGO", "DESCRIPCION",
                        "CANTIDAD", "S_MIN", "PEDIR"]]

    def process_all(self, path_med, path_acc, path_bal,
                    family_filter="Todas", params=None) -> pd.DataFrame:
        params = params or DEFAULT_PARAMS
        df     = self._full_calc(path_med, path_acc, path_bal, params)
        df     = self._apply_active_filter(df)
        return self._apply_filter(df, family_filter)

    def to_excel_bytes(self, df: pd.DataFrame) -> bytes:
        out = io.BytesIO()
        exp = df[["CODIGO", "DESCRIPCION", "PEDIR"]].copy()
        exp.columns = ["CÓDIGO", "DESCRIPCIÓN", "A PEDIR"]
        with pd.ExcelWriter(out, engine="openpyxl") as w:
            exp.to_excel(w, index=False, sheet_name="Pedidos")
            ws = w.sheets["Pedidos"]
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 55
            ws.column_dimensions["C"].width = 12
        out.seek(0)
        return out.read()

    def to_status_excel_bytes(self, path_med, path_acc, path_bal,
                               family_filter="Todas", params=None) -> bytes:
        params = params or DEFAULT_PARAMS
        df     = self._full_calc(path_med, path_acc, path_bal, params)
        df     = self._apply_active_filter(df)
        df     = self._apply_filter(df, family_filter)

        conditions = [
            df["CANTIDAD"] <= 0,
            df["CANTIDAD"] < df["S_MIN"],
            (df["S_MIN"] > 0) & (df["CANTIDAD"] >= 3 * df["S_MIN"]),
        ]
        df["STATUS"]    = np.select(conditions,
                                     [STATUS_SIN_STOCK, STATUS_FALTANTE, STATUS_SOBRESTOCK],
                                     default=STATUS_NORMAL)
        df["ARTÍCULO"]  = df["CODIGO"].astype(str) + " — " + df["DESCRIPCION"].astype(str)
        df["DIFERENCIA"] = (df["CANTIDAD"] - df["S_MIN"]).astype(int)

        export = df[["ARTÍCULO", "CANTIDAD", "S_MIN", "DIFERENCIA", "STATUS"]].copy()
        export.columns = ["Artículo", "Stock Actual", "Stock Mínimo", "Diferencia", "Status"]
        order_map = {STATUS_SIN_STOCK: 0, STATUS_FALTANTE: 1, STATUS_NORMAL: 2, STATUS_SOBRESTOCK: 3}
        export["_ord"] = export["Status"].map(order_map).fillna(2)
        export = (export.sort_values(["_ord", "Diferencia"])
                  .drop(columns=["_ord"]).reset_index(drop=True))

        out = io.BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as w:
            export.to_excel(w, index=False, sheet_name="Status Stock")
            ws = w.sheets["Status Stock"]
            ws.column_dimensions["A"].width = 55
            for col in ["B", "C", "D", "E"]:
                ws.column_dimensions[col].width = 16
            for row_idx, row_data in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=0):
                hex_c = STATUS_COLORS.get(
                    export.iloc[row_idx]["Status"] if row_idx < len(export) else STATUS_NORMAL,
                    "FFFFFF")
                fill = PatternFill("solid", fgColor=hex_c)
                for cell in row_data:
                    cell.fill = fill
            for cell in ws[1]:
                cell.font = Font(bold=True)
        out.seek(0)
        return out.read()

    # ══════════════════════════════════════════════════════════════════════════
    # FILTRO DE ARTÍCULOS ACTIVOS
    # ══════════════════════════════════════════════════════════════════════════

    def _load_active_codes(self) -> Optional[set]:
        if self._active_codes_cache is not None:
            return self._active_codes_cache
        if not self.forecast_path.exists():
            return None
        try:
            df = pd.read_excel(self.forecast_path, engine="openpyxl", dtype=str)
            df.columns = [c.strip() for c in df.columns]
            col_map: dict = {}
            for c in df.columns:
                cn = _strip_accents(c).lower()
                if "codigo" in cn:       col_map[c] = "Codigo"
                elif "estado" in cn:     col_map[c] = "Estado"
                elif "pronostico" in cn: col_map[c] = "Pronostico"
            df = df.rename(columns=col_map)
            if "Codigo" not in df.columns:
                return None
            df["Codigo"]     = df["Codigo"].astype(str).str.strip().str.lstrip("0").str.strip()
            df["Pronostico"] = pd.to_numeric(
                df.get("Pronostico", pd.Series(["0"] * len(df))), errors="coerce").fillna(0)
            fc_sum = df.groupby("Codigo")["Pronostico"].sum().reset_index()
            fc_sum.columns = ["Codigo", "FcTotal"]
            if "Estado" in df.columns:
                est = (df[["Codigo", "Estado"]].drop_duplicates("Codigo")
                       .set_index("Codigo")["Estado"].str.strip().str.upper())
                fc_sum["Estado"] = fc_sum["Codigo"].map(est).fillna("")
                mask = (fc_sum["Estado"] == "ACTIVO") | (fc_sum["FcTotal"] > 0)
            else:
                mask = fc_sum["FcTotal"] > 0
            active = set(fc_sum.loc[mask, "Codigo"].tolist())
            print(f"[FILTER] Artículos activos: {len(active)}", flush=True)
            self._active_codes_cache = active
            return active
        except Exception as exc:
            print(f"[FILTER] Warn: {exc}", flush=True)
            return None

    def _apply_active_filter(self, df: pd.DataFrame) -> pd.DataFrame:
        active = self._load_active_codes()
        if active is None:
            return df
        key      = df["CODIGO"].astype(str).str.strip().str.lstrip("0").str.strip()
        filtered = df[key.isin(active)].reset_index(drop=True)
        removed  = len(df) - len(filtered)
        if removed:
            print(f"[FILTER] Stock muerto excluido: {removed} artículos", flush=True)
        return filtered

    # ══════════════════════════════════════════════════════════════════════════
    # LÓGICA INTERNA
    # ══════════════════════════════════════════════════════════════════════════

    def _full_calc(self, path_med, path_acc, path_bal,
                   params, debug_dir=None) -> pd.DataFrame:
        master   = self._load_master()
        forecast = self._load_forecast()
        stock    = self._load_all_stocks(path_med, path_acc, path_bal)
        merged   = self._merge(master, stock, forecast)
        return self._calculate(merged, params, debug_dir)

    def _apply_filter(self, df: pd.DataFrame, family_filter: str) -> pd.DataFrame:
        if family_filter and family_filter.upper() != "TODAS":
            df = df[df["FAMILIA"].str.upper() == family_filter.upper()].reset_index(drop=True)
        return df

    def _load_master(self) -> pd.DataFrame:
        if self._master_cache is not None:
            return self._master_cache.copy()
        df = pd.read_excel(self.xlsm_path, sheet_name=XLSM_SHEET,
                           engine="openpyxl", header=0, usecols=XLSM_COLS_IDX)
        df.columns = XLSM_COLS_NAMES
        for col in ("CODIGO", "DESCRIPCION", "FAMILIA", "PROVEEDOR"):
            df[col] = df[col].fillna("").astype(str).str.strip()
        df["_PROV_U"] = df["PROVEEDOR"].str.upper()
        df = df[~df["_PROV_U"].isin(PROV_EXCLUDE)]
        df = df[df["PROVEEDOR"].str.len() >= 2]
        df = df[df["CODIGO"] != ""]
        df = df.drop(columns=["_PROV_U"])
        df["FAMILIA"]    = df["FAMILIA"].replace({"": "SIN FAMILIA", "0": "SIN FAMILIA"})
        df["COMPRA_MIN"] = pd.to_numeric(df["COMPRA_MIN"], errors="coerce").fillna(1).clip(lower=1)
        df["LOTE"]       = pd.to_numeric(df["LOTE"],       errors="coerce").fillna(1).clip(lower=1)
        self._master_cache = df.copy()
        print(f"[PROC] Maestro cargado: {len(df):,} artículos", flush=True)
        return df

    def _load_forecast(self) -> pd.DataFrame:
        if not self.forecast_path.exists():
            return pd.DataFrame(columns=["CODIGO", "DEMANDA_APLICABLE"])
        df = pd.read_excel(self.forecast_path, engine="openpyxl", dtype=str)
        df.columns = [c.strip() for c in df.columns]
        col_map: dict = {}
        for c in df.columns:
            cn = _strip_accents(c).lower()
            if "codigo" in cn:                           col_map[c] = "Codigo"
            elif "pronostico" in cn or "forecast" in cn: col_map[c] = "Pronostico"
            elif cn in ("mes", "month"):                 col_map[c] = "Mes"
            elif cn in ("año", "anio", "year", "ano"):   col_map[c] = "Anio"
        df = df.rename(columns=col_map)
        req = {"Codigo", "Pronostico", "Mes", "Anio"}
        if not req.issubset(df.columns):
            return pd.DataFrame(columns=["CODIGO", "DEMANDA_APLICABLE"])
        df["Codigo"]     = df["Codigo"].astype(str).str.strip().str.lstrip("0").str.strip()
        df["Pronostico"] = pd.to_numeric(df["Pronostico"], errors="coerce").fillna(0)
        df["Mes"]        = pd.to_numeric(df["Mes"],        errors="coerce").fillna(0).astype(int)
        df["Anio"]       = pd.to_numeric(df["Anio"],       errors="coerce").fillna(0).astype(int)
        today = date.today()
        if today.day <= 15:
            target_mes, target_anio = today.month, today.year
        else:
            target_mes  = 1 if today.month == 12 else today.month + 1
            target_anio = today.year + 1 if today.month == 12 else today.year
        print(f"[FORECAST] Día {today.day} → Mes={target_mes}, Año={target_anio}", flush=True)
        period = df[(df["Mes"] == target_mes) & (df["Anio"] == target_anio)].copy()
        period = period[["Codigo", "Pronostico"]].rename(
            columns={"Pronostico": "DEMANDA_APLICABLE", "Codigo": "CODIGO"})
        print(f"[FORECAST] {len(period)} registros período activo.", flush=True)
        return period

    def _load_all_stocks(self, path_med, path_acc, path_bal) -> pd.DataFrame:
        frames = []
        for path, label in [(path_med, "Medicamentos"),
                             (path_acc, "Accesorios"),
                             (path_bal, "Balanceados")]:
            frames.append(self._read_xls(Path(path), label))
        combined = pd.concat(frames, ignore_index=True)
        combined = combined.drop_duplicates(subset=["CODIGO"], keep="first")
        return combined[["CODIGO", "CANTIDAD"]]

    def _read_xls(self, path: Path, label: str) -> pd.DataFrame:
        ext = path.suffix.lower()
        print(f"[XLRD] Leyendo '{label}'…", flush=True)
        try:
            engine = "xlrd" if ext == ".xls" else "openpyxl"
            df     = pd.read_excel(path, engine=engine, header=0, dtype=str)
            return self._normalize_stock(df)
        except Exception as exc:
            raise ValueError(f"Error leyendo '{label}' ({path.name}): {exc}") from exc

    def _normalize_stock(self, df: pd.DataFrame) -> pd.DataFrame:
        col_map = self._build_col_map(df.columns)
        df = df.rename(columns=col_map)
        for col in ("CODIGO", "DESCRIPCION", "CANTIDAD"):
            if col not in df.columns:
                df[col] = ""
        df["CODIGO"]   = df["CODIGO"].fillna("").astype(str).str.strip()
        df["CANTIDAD"] = pd.to_numeric(df["CANTIDAD"], errors="coerce").fillna(0)
        return df[df["CODIGO"] != ""].reset_index(drop=True)

    def _build_col_map(self, columns) -> dict:
        found   = {k: False for k in ("CODIGO", "DESCRIPCION", "CANTIDAD", "MINIMO_REP", "DIFERENCIA")}
        mapping: dict = {}
        for col in columns:
            n = _strip_accents(str(col))
            if not found["CODIGO"] and n.startswith("artic") and "desc" not in n:
                mapping[col] = "CODIGO"; found["CODIGO"] = True
            elif not found["DESCRIPCION"] and (
                ("artic" in n and "desc" in n) or ("artic" in n and found["CODIGO"])
            ):
                mapping[col] = "DESCRIPCION"; found["DESCRIPCION"] = True
            elif not found["CANTIDAD"] and "cantidad" in n:
                mapping[col] = "CANTIDAD"; found["CANTIDAD"] = True
            elif not found["MINIMO_REP"] and ("minimo" in n or n.startswith("min")):
                mapping[col] = "MINIMO_REP"; found["MINIMO_REP"] = True
            elif not found["DIFERENCIA"] and "diferencia" in n:
                mapping[col] = "DIFERENCIA"; found["DIFERENCIA"] = True
        if not found["CODIGO"] or not found["CANTIDAD"]:
            pos = {0: "CODIGO", 1: "DESCRIPCION", 2: "CANTIDAD", 3: "MINIMO_REP", 4: "DIFERENCIA"}
            for idx, col in enumerate(columns):
                if idx in pos and col not in mapping:
                    mapping[col] = pos[idx]
        return mapping

    def _merge(self, master, stock, forecast) -> pd.DataFrame:
        master = master.copy(); stock = stock.copy()
        master["_KEY"] = master["CODIGO"].astype(str).str.strip().str.lstrip("0").str.strip()
        stock["_KEY"]  = stock["CODIGO"].astype(str).str.strip().str.lstrip("0").str.strip()
        merged = master.merge(stock[["_KEY", "CANTIDAD"]], on="_KEY", how="left")
        merged["CANTIDAD"] = pd.to_numeric(merged["CANTIDAD"], errors="coerce").fillna(0)
        if len(forecast) > 0:
            forecast = forecast.copy()
            forecast["_FKEY"] = forecast["CODIGO"].astype(str).str.strip().str.lstrip("0").str.strip()
            merged = merged.merge(
                forecast[["_FKEY", "DEMANDA_APLICABLE"]].rename(columns={"_FKEY": "_KEY"}),
                on="_KEY", how="left")
        else:
            merged["DEMANDA_APLICABLE"] = 0.0
        merged["DEMANDA_APLICABLE"] = pd.to_numeric(
            merged["DEMANDA_APLICABLE"], errors="coerce").fillna(0)
        cov = (merged["CANTIDAD"] != 0).sum()
        print(f"[PROC] Cruce: {len(merged):,} arts, {cov:,} con stock", flush=True)
        return merged

    def _calculate(self, df: pd.DataFrame, params: dict, debug_dir=None) -> pd.DataFrame:
        df = df.copy()
        df["COMPRA_MIN"]        = pd.to_numeric(df["COMPRA_MIN"],        errors="coerce").fillna(1).clip(lower=1)
        df["LOTE"]              = pd.to_numeric(df["LOTE"],              errors="coerce").fillna(1).clip(lower=1)
        df["DEMANDA_APLICABLE"] = pd.to_numeric(df["DEMANDA_APLICABLE"], errors="coerce").fillna(0)
        df["CANTIDAD"] = pd.to_numeric(df["CANTIDAD"], errors="coerce").fillna(0).clip(lower=0)
        df["S_MIN"] = 0.0
        for fam_key, fam_p in params.items():
            lt   = float(fam_p.get("lead_time", 2))
            seg  = fam_p.get("stock_seguridad", "2 semanas")
            mul  = SAFETY_MULTIPLIERS.get(seg, 0.5)
            mask = df["FAMILIA"].str.upper() == fam_key.upper()
            dd   = df.loc[mask, "DEMANDA_APLICABLE"] / 30.0
            df.loc[mask, "S_MIN"] = np.ceil(
                df.loc[mask, "DEMANDA_APLICABLE"] * mul + dd * lt).astype(float)
            print(f"[CALC] {fam_key}: {mask.sum()} arts, LT={lt}d, seg=×{mul}", flush=True)
        df["S_MIN"] = df["S_MIN"].clip(lower=0)
        no_pedir = (df["CANTIDAD"] >= df["S_MIN"]) | (df["S_MIN"] <= 0)
        needs    = df["S_MIN"] - df["CANTIDAD"]
        pedido   = np.maximum(df["COMPRA_MIN"], np.ceil(needs / df["LOTE"]) * df["LOTE"])
        df["PEDIR"] = np.where(no_pedir, 0, pedido).astype(int)
        if debug_dir:
            cols = ["CODIGO", "DESCRIPCION", "FAMILIA", "CANTIDAD",
                    "DEMANDA_APLICABLE", "S_MIN", "COMPRA_MIN", "LOTE", "PEDIR"]
            p = os.path.join(debug_dir, "debug_matematica.csv")
            df[cols].to_csv(p, index=False, encoding="utf-8-sig", sep=";")
        return df


# ═════════════════════════════════════════════════════════════════════════════
# ForecastEngine  —  Motor Holt-Winters AMORTIGUADO
# ═════════════════════════════════════════════════════════════════════════════
class ForecastEngine:

    # ── Caché en memoria ─────────────────────────────────────────────────────
    # Cada entrada se invalida cuando se sube un nuevo archivo al servidor.
    # La clave de invalidación es el path como string; si cambia el path
    # (o se llama a invalidate_caches()) el DataFrame se re-lee del disco.
    _cache_ventas:        Optional[pd.DataFrame] = None
    _cache_ventas_key:    str = ""

    _cache_compras:       Optional[pd.DataFrame] = None
    _cache_compras_key:   str = ""

    _cache_forecast_raw:  Optional[pd.DataFrame] = None
    _cache_forecast_key:  str = ""

    _cache_maestro_raw:   Optional[pd.DataFrame] = None
    _cache_maestro_key:   str = ""

    @classmethod
    def invalidate_caches(cls) -> None:
        """Llamar desde main.py cada vez que se sube o sobreescribe un archivo."""
        cls._cache_ventas       = None;  cls._cache_ventas_key   = ""
        cls._cache_compras      = None;  cls._cache_compras_key  = ""
        cls._cache_forecast_raw = None;  cls._cache_forecast_key = ""
        cls._cache_maestro_raw  = None;  cls._cache_maestro_key  = ""
        print("[FC-CACHE] Cachés invalidadas.", flush=True)

    # ── Utilidad Parquet ──────────────────────────────────────────────────────

    @staticmethod
    def _parquet_path(xlsx_path: Path) -> Path:
        """Devuelve la ruta .parquet paralela al xlsx (misma carpeta, mismo stem)."""
        return xlsx_path.with_suffix(".parquet")

    @staticmethod
    def _parquet_is_fresh(xlsx_path: Path, pq_path: Path) -> bool:
        """True si el .parquet existe y es más nuevo que el .xlsx fuente."""
        if not pq_path.exists():
            return False
        return pq_path.stat().st_mtime >= xlsx_path.stat().st_mtime

    @classmethod
    def _save_parquet(cls, df: pd.DataFrame, pq_path: Path) -> None:
        """
        Guarda df como Parquet de forma defensiva:
        - Convierte columnas con tipos mixtos o object problemáticos a string
          para que pyarrow no rechace el schema.
        - Si falla, loguea el error sin propagar (el Parquet es solo una
          optimización, nunca debe romper el flujo principal).
        """
        try:
            df_save = df.copy()
            for col in df_save.columns:
                if df_save[col].dtype == object:
                    # Forzar a string; NaN → cadena vacía
                    df_save[col] = df_save[col].fillna("").astype(str)
            df_save.to_parquet(pq_path, index=False, engine="pyarrow",
                               compression="snappy")
            print(f"[PARQUET] Guardado: {pq_path.name}", flush=True)
        except Exception as exc:
            print(f"[PARQUET] No se pudo guardar {pq_path.name}: {exc}", flush=True)

    # ── Helpers de carga con caché RAM + caché disco (Parquet) ────────────────

    @classmethod
    def _get_ventas(cls, path: Path) -> pd.DataFrame:
        key = str(path)
        # 1. Caché en RAM
        if cls._cache_ventas is not None and cls._cache_ventas_key == key:
            print("[FC-CACHE] Ventas → RAM.", flush=True)
            return cls._cache_ventas.copy()
        # 2. Caché en disco (Parquet)
        pq = cls._parquet_path(path)
        if cls._parquet_is_fresh(path, pq):
            try:
                df = pd.read_parquet(pq, engine="pyarrow")
                print(f"[PARQUET] Ventas → {pq.name} ({len(df):,} filas).", flush=True)
                cls._cache_ventas = df; cls._cache_ventas_key = key
                return df.copy()
            except Exception as exc:
                print(f"[PARQUET] Falló leer {pq.name}: {exc}. Leyendo xlsx.", flush=True)
        # 3. Disco — xlsx
        ext = path.suffix.lower()
        engine = "xlrd" if ext == ".xls" else "openpyxl"
        df = pd.read_excel(path, engine=engine)
        cls._save_parquet(df, pq)
        cls._cache_ventas = df; cls._cache_ventas_key = key
        print(f"[FC-CACHE] Ventas → xlsx ({len(df):,} filas).", flush=True)
        return df.copy()

    @classmethod
    def _get_compras(cls, path: Path) -> pd.DataFrame:
        """Caché de Compras — compartida con Analytics."""
        key = str(path)
        if cls._cache_compras is not None and cls._cache_compras_key == key:
            print("[FC-CACHE] Compras → RAM.", flush=True)
            return cls._cache_compras.copy()
        pq = cls._parquet_path(path)
        if cls._parquet_is_fresh(path, pq):
            try:
                df = pd.read_parquet(pq, engine="pyarrow")
                print(f"[PARQUET] Compras → {pq.name} ({len(df):,} filas).", flush=True)
                cls._cache_compras = df; cls._cache_compras_key = key
                return df.copy()
            except Exception as exc:
                print(f"[PARQUET] Falló leer {pq.name}: {exc}. Leyendo xlsx.", flush=True)
        ext = path.suffix.lower()
        engine = "xlrd" if ext == ".xls" else "openpyxl"
        df = pd.read_excel(path, engine=engine)
        cls._save_parquet(df, pq)
        cls._cache_compras = df; cls._cache_compras_key = key
        print(f"[FC-CACHE] Compras → xlsx ({len(df):,} filas).", flush=True)
        return df.copy()

    @classmethod
    def _get_forecast_raw(cls, path: Path) -> pd.DataFrame:
        key = str(path)
        if cls._cache_forecast_raw is not None and cls._cache_forecast_key == key:
            print("[FC-CACHE] Forecast → RAM.", flush=True)
            return cls._cache_forecast_raw.copy()
        pq = cls._parquet_path(path)
        if cls._parquet_is_fresh(path, pq):
            try:
                # Forecast se lee como dtype=str; guardar y leer preservando strings
                df = pd.read_parquet(pq, engine="pyarrow")
                # Re-forzar todo a str para mantener contrato
                df = df.astype(str)
                print(f"[PARQUET] Forecast → {pq.name} ({len(df):,} filas).", flush=True)
                cls._cache_forecast_raw = df; cls._cache_forecast_key = key
                return df.copy()
            except Exception as exc:
                print(f"[PARQUET] Falló leer {pq.name}: {exc}. Leyendo xlsx.", flush=True)
        df = pd.read_excel(path, engine="openpyxl", dtype=str)
        cls._save_parquet(df, pq)
        cls._cache_forecast_raw = df; cls._cache_forecast_key = key
        print(f"[FC-CACHE] Forecast → xlsx ({len(df):,} filas).", flush=True)
        return df.copy()

    @classmethod
    def _get_maestro_raw(cls, path: Path) -> pd.DataFrame:
        key = str(path)
        if cls._cache_maestro_raw is not None and cls._cache_maestro_key == key:
            print("[FC-CACHE] Maestro → RAM.", flush=True)
            return cls._cache_maestro_raw.copy()
        pq = cls._parquet_path(path)
        if cls._parquet_is_fresh(path, pq):
            try:
                df = pd.read_parquet(pq, engine="pyarrow")
                df = df.astype(str)
                print(f"[PARQUET] Maestro → {pq.name} ({len(df):,} filas).", flush=True)
                cls._cache_maestro_raw = df; cls._cache_maestro_key = key
                return df.copy()
            except Exception as exc:
                print(f"[PARQUET] Falló leer {pq.name}: {exc}. Leyendo xlsx.", flush=True)
        df = pd.read_excel(path, engine="openpyxl", dtype=str)
        cls._save_parquet(df, pq)
        cls._cache_maestro_raw = df; cls._cache_maestro_key = key
        print(f"[FC-CACHE] Maestro → xlsx ({len(df):,} filas).", flush=True)
        return df.copy()

    @staticmethod
    def _detect_ventas_cols(df: pd.DataFrame) -> dict:
        cols = [_strip_accents(str(c)).lower() for c in df.columns]
        orig = list(df.columns)
        def _find(*pats):
            for p in pats:
                for i, c in enumerate(cols):
                    if p in c:
                        return orig[i]
            return None
        return {
            "fecha":   _find("fecha", "emisi"),
            "cantidad":_find("cantidad"),
            "familia": _find("familia", "categor"),
            "articulo":next((orig[i] for i, c in enumerate(cols)
                             if "articulo" in c and "desc" not in c), None),
            "descrip": next((orig[i] for i, c in enumerate(cols)
                             if "articulo" in c and "desc" in c), _find("descrip")),
            "cliente": _find("denominaci", "razon", "cliente"),
        }

    @staticmethod
    def _map_maestro_cols(df: pd.DataFrame) -> pd.DataFrame:
        """
        v4.1 — Detecta y renombra columnas clave del Maestro de Productos.
        Prioridad: nombres exactos normalizados primero, heurísticas solo como fallback.
        Columnas canónicas resultantes:
          CodM      — código de artículo
          DescM     — descripción del artículo
          FamiliaM  — familia (ACCESORIOS / BALANCEADOS / MEDICAMENTOS FARMACIA)
          Categoria — subgrupo / descripcion2 / categoría de producto
        """
        # Mapa inverso: nombre_normalizado → columna_original
        norm2orig = {_strip_accents(c).lower().strip(): c for c in df.columns}
        cm: dict = {}

        # ── CodM (código de artículo) ─────────────────────────────────────────
        _COD_EXACT = ["articulo", "codigo articulo", "cod. articulo", "cod articulo",
                      "cod. art.", "cod.art.", "art. codigo", "codigo"]
        for name in _COD_EXACT:
            if name in norm2orig and "CodM" not in cm.values():
                cm[norm2orig[name]] = "CodM"
        if "CodM" not in cm.values():
            for orig in df.columns:
                cn = _strip_accents(orig).lower()
                if "articulo" in cn and "desc" not in cn and "CodM" not in cm.values():
                    cm[orig] = "CodM"

        # ── DescM (descripción del artículo) ──────────────────────────────────
        _DESC_EXACT = ["articulo descripcion", "desc. articulos", "desc articulos",
                       "descripcion articulo", "nombre articulo", "art. descripcion",
                       "desc. art."]
        for name in _DESC_EXACT:
            if name in norm2orig and "DescM" not in cm.values():
                cm[norm2orig[name]] = "DescM"
        if "DescM" not in cm.values():
            for orig in df.columns:
                cn = _strip_accents(orig).lower()
                if "articulo" in cn and "desc" in cn and "DescM" not in cm.values():
                    cm[orig] = "DescM"

        # ── FamiliaM (familia del negocio) ────────────────────────────────────
        _FAM_EXACT = ["familia", "familia articulo", "desc. familia",
                      "descripcion familia", "grupo principal", "grupo familia"]
        for name in _FAM_EXACT:
            if name in norm2orig and "FamiliaM" not in cm.values():
                cm[norm2orig[name]] = "FamiliaM"
        if "FamiliaM" not in cm.values():
            # Fallback 1 — nombre contiene "familia"
            for orig in df.columns:
                cn = _strip_accents(orig).lower()
                if "familia" in cn and "FamiliaM" not in cm.values():
                    cm[orig] = "FamiliaM"

        if "FamiliaM" not in cm.values():
            # Fallback 2 — detección por contenido: columna cuyos valores únicos
            # son un subconjunto de las 3 familias conocidas del negocio.
            # Cubre Maestros con columnas llamadas "Descripción", "Rubro", etc.
            _FAMS_KNOWN = {"ACCESORIOS", "BALANCEADOS", "MEDICAMENTOS FARMACIA"}
            _mapped_now = set(cm.keys())
            for orig in df.columns:
                if orig in _mapped_now:
                    continue
                try:
                    vals = set(
                        df[orig].dropna()
                               .astype(str).str.strip().str.upper()
                               .unique()
                    )
                    # La columna debe tener al menos 1 valor conocido y ninguno desconocido
                    if vals and vals.issubset(_FAMS_KNOWN):
                        cm[orig] = "FamiliaM"
                        print(
                            f"[MAESTRO] FamiliaM detectada por contenido: "
                            f"col='{orig}' vals={vals}",
                            flush=True,
                        )
                        break
                except Exception:
                    pass

        # ── Categoria (subgrupo / descripcion2) ───────────────────────────────
        _CAT_EXACT = ["descripcion2", "desc2", "descripcion 2", "subgrupo",
                      "sub grupo", "sub-grupo", "categoria", "category",
                      "grupo secundario", "descripcion grupo", "desc. grupo",
                      "desc grupo", "grupo articulo", "tipo articulo"]
        for name in _CAT_EXACT:
            if name in norm2orig and "Categoria" not in cm.values():
                cm[norm2orig[name]] = "Categoria"
        if "Categoria" not in cm.values():
            for orig in df.columns:
                cn = _strip_accents(orig).lower()
                if "Categoria" in cm.values():
                    break
                if (("grupo" in cn and "familia" not in cn) or
                        ("categ" in cn) or
                        (cn.endswith("2") and "desc" in cn)):
                    cm[orig] = "Categoria"

        df = df.rename(columns=cm)
        print(f"[MAESTRO] Cols detectadas: { {v: k for k, v in cm.items()} }", flush=True)
        return df

    @staticmethod
    def _construir_serie_mensual(df_art, fecha_col, qty_col) -> pd.Series:
        df2 = df_art.copy()
        df2["_mes"] = pd.to_datetime(df2[fecha_col]).dt.to_period("M").dt.to_timestamp()
        mensual = df2.groupby("_mes")[qty_col].sum()
        if mensual.empty:
            return pd.Series([], dtype=float)
        full = pd.date_range(start=mensual.index.min(),
                              end=mensual.index.max(), freq="MS")
        return mensual.reindex(full, fill_value=0)

    @staticmethod
    def _calcular_forecast_estacional(serie: pd.Series) -> list:
        """
        v3.5 — Holt-Winters AMORTIGUADO (Damped Trend).
        Parámetros exactos: damped_trend=True, smoothing_level=0.4,
          smoothing_trend=0.2, damping_trend=0.85
        Evita proyecciones explosivas al amortizar la tendencia.
        """
        if not _HAS_STATSMODELS:
            raise RuntimeError("statsmodels no instalado.")
        try:
            if len(serie) < 24:
                idx  = pd.date_range(end=serie.index[-1], periods=24, freq="MS")
                serie = serie.reindex(idx, fill_value=0)

            # v3.5 — Damped Trend Holt-Winters
            model = _HW(
                serie,
                trend="add",
                seasonal="add",
                seasonal_periods=FC_SEASONAL_PERIODS,
                damped_trend=True,
            )
            fit = model.fit(
                smoothing_level=FC_SMOOTHING_LEVEL,
                smoothing_trend=FC_SMOOTHING_TREND,
                damping_trend=FC_DAMPING_TREND,
                optimized=False,
            )
            pred = np.clip(fit.forecast(12).values, 0, None)

            # Fallback si todos los valores son iguales (serie plana)
            if len(set(np.round(pred, 2))) == 1:
                base = max(serie.tail(3).mean() if serie.tail(3).mean() > 0 else serie.mean(), 0.1)
                pred = np.array([base * (1 + i * 0.01) for i in range(1, 13)])

            return [max(0, round(float(v))) for v in pred]
        except Exception as exc:
            print(f"[FC] Fallback por excepción: {exc}", flush=True)
            fallback = max(0, round(float(serie.tail(6).mean())))
            return [fallback] * 12

    @staticmethod
    def _next_12_months_from_max(df_v: pd.DataFrame, fecha_col: str) -> list:
        max_date = pd.to_datetime(df_v[fecha_col], errors="coerce").max()
        if pd.isnull(max_date):
            max_date = pd.Timestamp.now()
        print(f"[FC] Fecha máxima historial: {max_date.strftime('%Y-%m')}", flush=True)
        start_m = max_date.month + 1 if max_date.month < 12 else 1
        start_y = max_date.year      if max_date.month < 12 else max_date.year + 1
        months  = []
        for i in range(12):
            off = start_m - 1 + i
            months.append((off % 12 + 1, start_y + off // 12))
        print(f"[FC] Período: {months[0]} → {months[-1]}", flush=True)
        return months

    # ── API pública ───────────────────────────────────────────────────────────

    @classmethod
    def run_and_save(cls, ventas_path: Path, maestro_path: Path,
                     forecast_output_path: Path) -> dict:
        if not _HAS_STATSMODELS:
            raise RuntimeError("statsmodels no instalado.")
        df_v = pd.read_excel(ventas_path,  engine="openpyxl")
        df_m = pd.read_excel(maestro_path, engine="openpyxl")
        df_m.columns = [str(c).strip() for c in df_m.columns]
        art_col = next((c for c in df_m.columns
                        if "articulo" in _strip_accents(str(c)).lower()
                        and "desc" not in _strip_accents(str(c)).lower()),
                       df_m.columns[0])
        desc_col = next((c for c in df_m.columns
                         if "desc" in _strip_accents(str(c)).lower()), None)
        df_m[art_col] = df_m[art_col].astype(str).str.strip()
        dc = cls._detect_ventas_cols(df_v)
        if not dc["fecha"] or not dc["cantidad"] or not dc["articulo"]:
            raise ValueError("No se detectaron columnas clave en Ventas.xlsx.")
        df_v[dc["articulo"]] = df_v[dc["articulo"]].astype(str).str.strip()
        df_v[dc["fecha"]]    = pd.to_datetime(df_v[dc["fecha"]], dayfirst=True, errors="coerce")
        df_v = df_v.dropna(subset=[dc["fecha"]])
        df_v[dc["cantidad"]] = pd.to_numeric(df_v[dc["cantidad"]], errors="coerce").fillna(0)
        df_v = df_v[df_v[dc["cantidad"]] > 0]
        if dc["familia"]:
            df_v[dc["familia"]] = df_v[dc["familia"]].astype(str).str.strip().str.upper()
            df_v = df_v[df_v[dc["familia"]].isin([f.upper() for f in FC_FAMILIAS])].copy()
        if df_v.empty:
            raise ValueError("Sin registros de ventas para las familias requeridas.")
        forecast_months = cls._next_12_months_from_max(df_v, dc["fecha"])
        fecha_corte = pd.Timestamp.now() - pd.DateOffset(months=FC_MESES_RECENCIA)
        resumen     = df_v.groupby(dc["articulo"]).agg(
            total=(dc["cantidad"], "sum"), ultima=(dc["fecha"], "max"))
        activos = resumen[(resumen["total"] >= FC_MIN_VENTAS) &
                          (resumen["ultima"] >= fecha_corte)].index
        def _info(cod):
            sub = df_v[df_v[dc["articulo"]] == cod]
            d   = str(sub[dc["descrip"]].iloc[-1] or "") if not sub.empty and dc.get("descrip") else ""
            f   = str(sub[dc["familia"]].iloc[-1] or "") if not sub.empty and dc.get("familia") else ""
            if not d and desc_col:
                row_m = df_m[df_m[art_col] == cod]
                d = str(row_m[desc_col].iloc[0] or "") if not row_m.empty else ""
            return d, f
        todos   = df_m[art_col].unique()
        rows: list = []
        print(f"[FC] Procesando {len(todos)} artículos (Holt-Winters Amortiguado)…", flush=True)
        for i, cod in enumerate(todos):
            if i % 100 == 0:
                print(f"[FC]   {i}/{len(todos)}", flush=True)
            es_act = cod in activos
            d, f   = _info(cod)
            if not es_act:
                vals = [0] * 12
            else:
                df_a  = df_v[df_v[dc["articulo"]] == cod]
                serie = cls._construir_serie_mensual(df_a, dc["fecha"], dc["cantidad"])
                vals  = cls._calcular_forecast_estacional(serie) \
                    if not serie.empty and len(serie) >= 2 else [0] * 12
            for (mes, anio), val in zip(forecast_months, vals):
                rows.append({
                    "Codigo":     cod, "Producto": d, "Familia": f,
                    "Mes": mes, "Año": anio, "Pronostico": int(val),
                    "Estado": "ACTIVO" if es_act else "INACTIVO",
                })
        pd.DataFrame(rows).to_excel(forecast_output_path, index=False, engine="openpyxl")
        total = int(sum(r["Pronostico"] for r in rows))
        meses = [f"{m:02d}/{y}" for m, y in forecast_months]
        print(f"[FC] ✓ Total forecast: {total:,} uds", flush=True)
        return {
            "n_articulos":    len(todos),
            "n_activos":      len(activos),
            "n_inactivos":    len(todos) - len(activos),
            "total_forecast": total,
            "meses_forecast": meses,
            "periodo_inicio": meses[0],
            "periodo_fin":    meses[-1],
        }

    @classmethod
    def get_forecast_filtros(cls, forecast_path: Path,
                              maestro_path: Optional[Path] = None) -> dict:
        """
        v4.1 — Maestro de Productos es la fuente EXCLUSIVA de familias y categorías.
        Las 3 familias del negocio son la fuente de verdad (hardcodeadas).
        Los artículos provienen del Forecast (activos) y se enriquecen con el Maestro.
        """
        # Familias fijas del negocio — nunca cambian
        FAMILIAS_NEGOCIO = ["ACCESORIOS", "BALANCEADOS", "MEDICAMENTOS FARMACIA"]

        # Inicializar cats_by_fam con las 3 familias garantizadas (aunque no tengan categorías)
        cats_by_fam: dict = {fam: [] for fam in FAMILIAS_NEGOCIO}
        articulos:   list = []
        df_fc: Optional[pd.DataFrame] = None

        # ── Paso 1: Forecast.xlsx — solo para lista de artículos activos ──────
        if forecast_path and forecast_path.exists():
            try:
                df_fc = cls._get_forecast_raw(forecast_path)
                df_fc.columns = [c.strip() for c in df_fc.columns]
                cm_fc: dict = {}
                for c in df_fc.columns:
                    cn = _strip_accents(c).lower()
                    if "codigo" in cn:       cm_fc[c] = "Codigo"
                    elif "familia" in cn:    cm_fc[c] = "Familia"
                    elif "estado" in cn:     cm_fc[c] = "Estado"
                    elif "pronostico" in cn: cm_fc[c] = "Pronostico"
                    elif "producto" in cn:   cm_fc[c] = "Producto"
                df_fc = df_fc.rename(columns=cm_fc)
                if "Codigo" in df_fc.columns:
                    df_fc["Codigo"] = df_fc["Codigo"].astype(str).str.strip().str.lstrip("0")
                if "Codigo" in df_fc.columns:
                    desc_col_fc = "Producto" if "Producto" in df_fc.columns else None
                    fam_col_fc  = "Familia"  if "Familia"  in df_fc.columns else None
                    for _, row in df_fc.drop_duplicates("Codigo").iterrows():
                        cod  = str(row.get("Codigo", "")).strip()
                        desc = str(row.get(desc_col_fc, "") or "") if desc_col_fc else ""
                        fam  = str(row.get(fam_col_fc,  "") or "") if fam_col_fc  else ""
                        if cod:
                            articulos.append({
                                "codigo":      cod,
                                "descripcion": desc,
                                "familia":     fam,
                                "categoria":   "",   # se enriquece en el Paso 2
                            })
            except Exception as exc:
                print(f"[FILTROS] Warn forecast: {exc}", flush=True)

        # ── Paso 2: Maestro — FUENTE EXCLUSIVA de familias y categorías ───────
        if maestro_path and maestro_path.exists():
            try:
                df_m = cls._get_maestro_raw(maestro_path)
                df_m.columns = [c.strip() for c in df_m.columns]

                # Usar el mapper centralizado con precedencia por nombre exacto
                df_m = cls._map_maestro_cols(df_m)

                # Normalizar campos
                if "CodM" in df_m.columns:
                    df_m["CodM"] = df_m["CodM"].astype(str).str.strip().str.lstrip("0")
                if "FamiliaM" in df_m.columns:
                    df_m["FamiliaM"] = df_m["FamiliaM"].astype(str).str.strip().str.upper()
                if "Categoria" in df_m.columns:
                    df_m["Categoria"] = df_m["Categoria"].astype(str).str.strip().str.upper()

                # ── Lookup código → {familia, categoria} desde el Maestro ──────
                lookup: dict = {}
                if "CodM" in df_m.columns:
                    for _, row in df_m.iterrows():
                        cod_m = str(row.get("CodM", "")).strip()
                        if not cod_m:
                            continue
                        fam_m = str(row.get("FamiliaM", "") or "").strip() \
                            if "FamiliaM" in df_m.columns else ""
                        cat_m = str(row.get("Categoria", "") or "").strip() \
                            if "Categoria" in df_m.columns else ""
                        # Solo guardar si pertenece a una familia válida del negocio
                        if fam_m in FAMILIAS_NEGOCIO:
                            lookup[cod_m] = {"familia": fam_m, "categoria": cat_m}

                # ── Artículos desde Maestro si Forecast estaba vacío ──────────
                if not articulos and "CodM" in df_m.columns:
                    desc_col_m = "DescM" if "DescM" in df_m.columns else None
                    for _, row in df_m.drop_duplicates("CodM").iterrows():
                        cod  = str(row.get("CodM", "")).strip()
                        desc = str(row.get(desc_col_m, "") or "") if desc_col_m else ""
                        info = lookup.get(cod, {})
                        if cod:
                            articulos.append({
                                "codigo":      cod,
                                "descripcion": desc,
                                "familia":     info.get("familia", ""),
                                "categoria":   info.get("categoria", ""),
                            })
                else:
                    # Enriquecer artículos del Forecast con datos del Maestro
                    for art in articulos:
                        info = lookup.get(art["codigo"], {})
                        if info.get("categoria"):
                            art["categoria"] = info["categoria"]
                        if info.get("familia"):
                            # El Maestro siempre prevalece sobre el Forecast
                            art["familia"] = info["familia"]

                # ── Categorías por familia — EXCLUSIVAMENTE desde el Maestro ──
                if "CodM" in df_m.columns and "Categoria" in df_m.columns \
                   and "FamiliaM" in df_m.columns:
                    # Filtrar solo las 3 familias del negocio y categorías válidas
                    _INVALID_CATS = {"NAN", "NONE", "0", ""}
                    df_valid = df_m[
                        df_m["FamiliaM"].isin(FAMILIAS_NEGOCIO) &
                        df_m["Categoria"].notna() &
                        (df_m["Categoria"].str.strip() != "") &
                        (~df_m["Categoria"].str.upper().isin(_INVALID_CATS))
                    ]
                    for fam_val, grp in df_valid.groupby("FamiliaM"):
                        cats = sorted(
                            c for c in grp["Categoria"].astype(str).str.strip().unique()
                            if c and c.upper() not in _INVALID_CATS)
                        if cats:
                            cats_by_fam[str(fam_val).strip().upper()] = cats

                elif df_fc is not None and "Familia" in df_fc.columns \
                     and "Codigo" in df_fc.columns \
                     and "CodM" in df_m.columns \
                     and "Categoria" in df_m.columns:
                    # Fallback: cruzar Forecast + Maestro si el Maestro no tiene FamiliaM
                    merged = df_fc[["Codigo", "Familia"]].drop_duplicates().merge(
                        df_m[["CodM", "Categoria"]], left_on="Codigo", right_on="CodM",
                        how="left")
                    for fam, grp in merged.groupby("Familia"):
                        fam_u = str(fam).strip().upper()
                        if fam_u not in FAMILIAS_NEGOCIO:
                            continue
                        cats = sorted(
                            c for c in grp["Categoria"].dropna()
                                      .astype(str).str.strip().unique()
                            if c and c.upper() not in {"NAN", "NONE", "", "0"})
                        if cats:
                            cats_by_fam[fam_u] = cats

            except Exception as exc:
                print(f"[FILTROS] Warn maestro: {exc}", flush=True)

        print(
            f"[FILTROS] familias={FAMILIAS_NEGOCIO} "
            f"cats={ {k: len(v) for k, v in cats_by_fam.items()} } "
            f"arts={len(articulos)}",
            flush=True,
        )
        return {
            "familias":              FAMILIAS_NEGOCIO,
            "categorias_by_familia": cats_by_fam,
            "articulos":             articulos,
        }

    @classmethod
    def get_timeseries_for_chart(
        cls,
        ventas_path: Path,
        forecast_path: Path,
        maestro_path: Optional[Path] = None,
        familia: str   = "Todas",
        categoria: str = "Todas",   # parámetro mantenido por compatibilidad; ignorado
        articulo: str  = "",
    ) -> dict:
        """
        v4.2 — Cambios respecto a v4.1:
        1. Familia se resuelve SIEMPRE vía Maestro (FamiliaM), nunca por la
           columna Familia del Forecast.xlsx. Garantiza que histórico y
           pronóstico sumen exactamente los mismos artículos.
        2. `categoria` se ignora (eliminado de la UI en esta versión).
        3. Ventas, Forecast y Maestro se leen desde caché en memoria.
        """
        fam_upper = familia.strip().upper()
        cat_upper = categoria.strip().upper()
        art_q     = articulo.strip().lower()

        # ── Paso 1: cod_activos desde el Maestro ─────────────────────────────
        # Se activa si hay filtro de familia, categoría o artículo libre.
        cod_activos: Optional[set] = None
        if maestro_path and maestro_path.exists() and (
            fam_upper not in ("TODAS", "") or
            cat_upper not in ("TODAS", "") or
            art_q
        ):
            try:
                df_m = cls._get_maestro_raw(maestro_path)
                df_m.columns = [c.strip() for c in df_m.columns]
                df_m = cls._map_maestro_cols(df_m)
                if "CodM" in df_m.columns:
                    df_m["CodM"] = df_m["CodM"].astype(str).str.strip().str.lstrip("0")
                    mask = pd.Series([True] * len(df_m), index=df_m.index)
                    # Filtro de familia desde el Maestro
                    if "FamiliaM" in df_m.columns and fam_upper not in ("TODAS", ""):
                        df_m["FamiliaM"] = (df_m["FamiliaM"].astype(str)
                                            .str.strip().str.upper())
                        mask &= df_m["FamiliaM"] == fam_upper
                    # Filtro de categoría desde el Maestro (Req 1)
                    if "Categoria" in df_m.columns and cat_upper not in ("TODAS", ""):
                        df_m["Categoria"] = (df_m["Categoria"].astype(str)
                                             .str.strip().str.upper())
                        mask &= df_m["Categoria"] == cat_upper
                    # Filtro de artículo por texto libre
                    if art_q:
                        m_art = df_m["CodM"].str.lower().str.contains(art_q, na=False)
                        if "DescM" in df_m.columns:
                            m_art |= (df_m["DescM"].astype(str).str.lower()
                                      .str.contains(art_q, na=False))
                        mask &= m_art
                    cod_activos = set(df_m.loc[mask, "CodM"].tolist())
                    print(
                        f"[TS] Maestro filtrado (fam={fam_upper}, cat={cat_upper}) "
                        f"→ {len(cod_activos)} códigos activos",
                        flush=True,
                    )
            except Exception as exc:
                print(f"[TS] Warn maestro: {exc}", flush=True)

        # ── Paso 2: Forecast desde caché ─────────────────────────────────────
        fc_map:     dict = {}
        fc_cod_all: Optional[set] = None   # todos los códigos presentes en Forecast
        last_hist:  Optional[str] = None
        if forecast_path and forecast_path.exists():
            try:
                df_f = cls._get_forecast_raw(forecast_path)
                df_f.columns = [c.strip() for c in df_f.columns]
                cm: dict = {}
                for c in df_f.columns:
                    cn = _strip_accents(c).lower()
                    if "codigo" in cn:               cm[c] = "Codigo"
                    elif "pronostico" in cn:         cm[c] = "Pronostico"
                    elif cn in ("mes", "month"):     cm[c] = "Mes"
                    elif cn in ("año", "anio", "year", "ano"): cm[c] = "Anio"
                df_f = df_f.rename(columns=cm)
                df_f["Codigo"]     = df_f["Codigo"].astype(str).str.strip().str.lstrip("0")
                df_f["Pronostico"] = pd.to_numeric(
                    df_f.get("Pronostico", 0), errors="coerce").fillna(0)
                df_f["Mes"]  = pd.to_numeric(
                    df_f.get("Mes",  0), errors="coerce").fillna(0).astype(int)
                df_f["Anio"] = pd.to_numeric(
                    df_f.get("Anio", 0), errors="coerce").fillna(0).astype(int)

                fc_cod_all = set(df_f["Codigo"].unique())

                # Filtrar por cod_activos (resuelto desde el Maestro)
                if cod_activos is not None:
                    df_f = df_f[df_f["Codigo"].isin(cod_activos)]

                if "Mes" in df_f.columns and "Anio" in df_f.columns:
                    df_f["_ym"] = (
                        df_f["Anio"].astype(int).astype(str) + "-" +
                        df_f["Mes"].astype(int).apply(lambda m: f"{m:02d}"))
                    agg    = df_f.groupby("_ym")["Pronostico"].sum()
                    fc_map = {k: int(round(v)) for k, v in agg.items() if v > 0}
            except Exception as exc:
                print(f"[TS] Warn forecast: {exc}", flush=True)

        # ── Paso 3: Ventas históricas desde caché ─────────────────────────────
        hist_map: dict = {}
        if ventas_path and ventas_path.exists():
            try:
                df_v = cls._get_ventas(ventas_path)
                dc   = cls._detect_ventas_cols(df_v)
                if dc["fecha"] and dc["cantidad"]:
                    df_v[dc["fecha"]]    = pd.to_datetime(
                        df_v[dc["fecha"]], dayfirst=True, errors="coerce")
                    df_v[dc["cantidad"]] = pd.to_numeric(
                        df_v[dc["cantidad"]], errors="coerce").fillna(0)
                    df_v = df_v.dropna(subset=[dc["fecha"]])
                    df_v = df_v[df_v[dc["cantidad"]] > 0]

                    # Filtro de artículos por cod_activos (resuelto vía Maestro)
                    if dc["articulo"]:
                        df_v[dc["articulo"]] = (df_v[dc["articulo"]]
                                                .astype(str).str.strip().str.lstrip("0"))
                        allowed = cod_activos if cod_activos is not None else fc_cod_all
                        if allowed:
                            df_v = df_v[df_v[dc["articulo"]].isin(allowed)]

                    df_v["_ym"] = df_v[dc["fecha"]].dt.strftime("%Y-%m")
                    agg      = df_v.groupby("_ym")[dc["cantidad"]].sum()
                    hist_map = {k: int(round(v)) for k, v in agg.items()}
                    if hist_map:
                        last_hist = sorted(hist_map.keys())[-1]
            except Exception as exc:
                print(f"[TS] Warn ventas: {exc}", flush=True)

        all_keys = sorted(set(hist_map.keys()) | set(fc_map.keys()))
        if not all_keys:
            return {
                "familia": familia, "labels": [], "historico": [],
                "forecast": [], "stats": {
                    "total_historico": 0, "total_forecast": 0,
                    "meses_con_ventas": 0, "avg_historico": 0, "avg_forecast": 0,
                },
                "filtros_aplicados": {"familia": familia, "categoria": categoria, "articulo": articulo},
            }

        historico: list = []
        forecast:  list = []
        for ym in all_keys:
            is_h = ym in hist_map
            is_f = ym in fc_map
            historico.append(hist_map[ym] if is_h else None)
            if is_f and not is_h:
                forecast.append(fc_map[ym])
            elif is_f and ym == last_hist:
                forecast.append(fc_map[ym])
            else:
                forecast.append(None)

        total_historico  = sum(v for v in historico if v is not None)
        total_forecast   = sum(v for v in forecast  if v is not None)
        meses_con_ventas = sum(1 for v in historico if v is not None and v > 0)
        avg_historico    = round(total_historico / max(meses_con_ventas, 1))
        avg_forecast     = round(total_forecast  / 12)

        return {
            "familia":   familia,
            "labels":    all_keys,
            "historico": historico,
            "forecast":  forecast,
            "stats": {
                "total_historico":  total_historico,
                "total_forecast":   total_forecast,
                "meses_con_ventas": meses_con_ventas,
                "avg_historico":    avg_historico,
                "avg_forecast":     avg_forecast,
            },
            "filtros_aplicados": {"familia": familia, "categoria": categoria, "articulo": articulo},
        }


# ─── Utilidad global ──────────────────────────────────────────────────────────

def _strip_accents(text: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    ).lower().strip()